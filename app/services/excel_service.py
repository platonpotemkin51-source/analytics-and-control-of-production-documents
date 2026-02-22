from pathlib import Path

from openpyxl import Workbook, load_workbook

BASE = Path("generated_excels")
TEMPLATE_DIR = Path("app/excel/templates")
TEMPLATE_DIR.mkdir(parents=True, exist_ok=True)
BASE.mkdir(parents=True, exist_ok=True)


def _ensure_template(name: str) -> Path:
    path = TEMPLATE_DIR / name
    if not path.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "Shipment"
        ws["A1"] = "WoodFlow shipment document"
        wb.save(path)
    return path


def generate_documents(shipment_id: int, payload: dict) -> dict:
    output = {}
    for template_name, prefix in (("buyer_template.xlsx", "buyer"), ("customs_template.xlsx", "customs")):
        template_path = _ensure_template(template_name)
        wb = load_workbook(template_path)
        ws = wb.active
        ws["A2"] = f"Shipment ID: {shipment_id}"
        ws["A3"] = f"Contract: {payload['contract']}"
        ws["A4"] = f"Container: {payload['container_number']}"
        ws["A5"] = f"Total cost: {payload['total_cost']}"
        ws["A6"] = f"Batch IDs: {', '.join(map(str, payload['batch_ids']))}"
        file_path = BASE / f"{prefix}_shipment_{shipment_id}.xlsx"
        wb.save(file_path)
        output[prefix] = str(file_path)
    return output
